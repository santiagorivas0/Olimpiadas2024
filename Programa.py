"""
Programa.py es el punto de entrada de nuestro sistema:
- Presenta un menú interactivo hacia el usuario final.
- Permite gestionar Atletas, Disciplinas, Eventos y Participaciones.
- Utiliza las reglas de negocio definidas en Reglas.py para realizar las operaciones.
"""

#Obligatorio 2024 - Introducción a la Programación II / Dorrego,Rivas
from os import system, sys, path
from datetime import date
from Reglas import Olimpiada, Disciplina #Importa la Controladora (Olimpiada) y la clase Disciplina porque la necesitaremos.
import pickle
import Herramientas
from openpyxl import load_workbook
from matplotlib import pyplot as plt

juegos_olimpicos = Olimpiada()

def mostrar_menu():
    print("\nMenú de opciones")
    print("1 - Ingreso de Atletas")
    print("2 - Ingreso de Disciplinas desde Excel")
    print("3 - Asignación de Disciplina a Atletas")
    print("4 - Ingreso de Eventos")
    print("5 - Ingreso de Participaciones")
    print("6 - Cantidad de Atletas por Disciplina (Gráfica)")
    print("0 - Salir")

def solicitar_opcion():
    opcion = int(input("Ingrese una opción: "))
    while opcion not in [0,1,2,3,4,5,6]:
        print("La opción ingresada es inválida. Intente nuevamente.")
        opcion = int(input("Ingrese una opción: "))
    return opcion

#Serialización Binaria
def cargar_datos():
    global juegos_olimpicos
    ruta_archivo = sys.path[0]
    nom_archivo = "datos.bin"
    ruta_completa = path.join(ruta_archivo, nom_archivo)
    if path.isfile(ruta_completa):
        with open(ruta_completa, "rb") as stream:
            juegos_olimpicos = pickle.load(stream)

def guardar_datos():
    ruta_archivo = sys.path[0]
    nom_archivo = "datos.bin"
    ruta_completa = path.join(ruta_archivo, nom_archivo)
    with open(ruta_completa, "wb") as stream:
        pickle.dump(juegos_olimpicos, stream)

def ingresar_atleta():
    #Se valida el ID ingresado y existencia del Atleta.
    id_atleta = Herramientas.validar_codigo(input("Ingrese el ID del atleta (4 dígitos): "), 4)
    atleta_buscado = juegos_olimpicos.buscar_atleta(int(id_atleta))
    if atleta_buscado is None:
        nombre = input("Ingrese el nombre completo del atleta: ")
        sexo = input("Ingrese el sexo del atleta (M/F): ")
        pais = input("Ingrese el país del atleta: ")
        registro_ok = juegos_olimpicos.agregar_atleta(nombre, sexo, int(id_atleta), pais)
        if registro_ok:
            print("Atleta registrado exitosamente.")
        else:
            print("Ocurrió un problema al registrar el atleta.")
        #Muestra todos los Atletas registrados después de agregar el nuevo Atleta.
        print("\nAtletas registrados:")
        for atleta in juegos_olimpicos.lista_atletas:
            disciplina_nombre = atleta.disciplina.nombre if atleta.disciplina else "No Asignada"
            print(f"ID: {atleta.id_atleta} --- Nombre: {atleta.nombre_completo} --- Sexo: {atleta.sexo} --- País: {atleta.pais} --- Disciplina: {disciplina_nombre}")
    else:
        print("Ya existe un atleta con ese ID.")

def ingreso_disciplinas_desde_excel():
    #Lee las disciplinas desde un archivo Excel (disciplinas.xlsx) y las agrega al sistema.
    ruta_archivo = path.join(sys.path[0], "Disciplinas.xlsx")
    libro = load_workbook(ruta_archivo)
    hoja = libro["DisciplinasID"] #Especifica el nombre de la hoja.
    print("Disciplinas cargadas desde el archivo 'Disciplinas.xlsx':")
    for fila in hoja.iter_rows(min_row=2, values_only=True): #Omite la primera fila (títulos).
        id_disciplina, nombre_disciplina = fila
        if juegos_olimpicos.buscar_disciplina(id_disciplina) is None:
            juegos_olimpicos.lista_disciplinas.append(Disciplina(id_disciplina, nombre_disciplina))
            print(f"- ID: {id_disciplina}, Nombre: {nombre_disciplina}") #Muestra todas las disciplinas, solo agrega las que no están cargadas, omitirá ID 1 y 2 ya existentes en Reglas.py.
    print("Carga de disciplinas finalizada.")

def asignar_disciplina_atleta():
    #Muestra la lista de atletas registrados
    for atleta in juegos_olimpicos.lista_atletas:
        print(f"{atleta.id_atleta} - {atleta.nombre_completo}")
    #Solicita el ID del atleta y verifica su existencia
    id_atleta = Herramientas.solicitar_entero("Ingrese el ID del atleta: ")
    atleta = juegos_olimpicos.buscar_atleta(id_atleta)  #Busca el atleta para validar.

    if atleta:
        #Muestra la lista de disciplinas disponibles
        for disciplina in juegos_olimpicos.lista_disciplinas:
            print(f"{disciplina.id_disciplina} - {disciplina.nombre}")
        #Solicita el ID de la disciplina y verifica su existencia
        id_disciplina = Herramientas.solicitar_entero("Ingrese el ID de la disciplina: ")
        disciplina = juegos_olimpicos.buscar_disciplina(id_disciplina)  #Busca la disciplina para validar.

        if disciplina:
            #Si el atleta no tiene disciplina asignada, se asigna por primera vez
            if atleta.disciplina is None:
                atleta.disciplina = disciplina
                print("Disciplina asignada al atleta correctamente.")
            #Si el atleta tiene una disciplina distinta, se reasigna
            elif atleta.disciplina != disciplina:
                atleta.disciplina = disciplina
                print("Disciplina del atleta reasignada correctamente.")
            #Si la disciplina ingresada ya está asignada, no hay cambios
            else:
                print("El atleta ya tiene asignada esa disciplina.")
        else:
            print("Disciplina no encontrada.")
    else:
        print("Atleta no encontrado.")

def ingresar_evento():
    id_evento = Herramientas.validar_codigo(input("Ingrese el ID del evento (3 dígitos): "), 3)
    evento_buscado = juegos_olimpicos.buscar_evento(int(id_evento)) 
    
    if evento_buscado is None: #Busca que no exista evento para poder inciarlo.
        nombre_prueba = input("Ingrese el nombre de la prueba: ")
        
        #Muestra las disciplinas disponibles con sus IDs
        print("\nDisciplinas disponibles:")
        for disciplina in juegos_olimpicos.lista_disciplinas:
            print(f"ID: {disciplina.id_disciplina} --- Nombre: {disciplina.nombre}") #Recorre la lista de disciplinas e imprime el ID y el nombre.
        
        id_disciplina = Herramientas.validar_codigo(input("Ingrese el ID de la disciplina (1 dígito): "), 1)
        disciplina = juegos_olimpicos.buscar_disciplina(int(id_disciplina)) #Busca que exista el ID de la disciplina para posteriormete asignarle fecha de inicio y fin
        
        if disciplina:
            fecha_inicio = Herramientas.solicitar_fecha("Ingrese la fecha de inicio del evento")
            fecha_fin = Herramientas.solicitar_fecha("Ingrese la fecha de fin del evento")
            juegos_olimpicos.agregar_evento(int(id_evento),nombre_prueba,int(id_disciplina),fecha_inicio, fecha_fin)
            print("Evento registrado exitosamente.")
        else:
            print("La disciplina ingresada no existe.")
    else:
        print("Ya existe un evento con ese ID.")

def registrar_participacion():
    #Solicita la fecha del evento y filtra los eventos disponibles en esa fecha
    fecha = Herramientas.solicitar_fecha("Ingrese la fecha del evento")
    eventos_en_fecha = [evento for evento in juegos_olimpicos.lista_eventos if evento.fecha_inicio <= fecha <= evento.fecha_fin]
    
    if eventos_en_fecha:
        #Muestra los eventos disponibles en la fecha ingresada
        print("Eventos disponibles en la fecha:")
        for evento in eventos_en_fecha:
            print(f"{evento.id_evento} - {evento.nombre_prueba}")
        
        #Solicita el ID del evento y verifica su existencia
        id_evento = Herramientas.solicitar_entero("Ingrese el ID del evento: ")
        evento = juegos_olimpicos.buscar_evento(id_evento)

        if evento:
            #Solicita el ID del atleta y verifica su existencia
            id_atleta = Herramientas.solicitar_entero("Ingrese el ID del atleta: ")
            atleta = juegos_olimpicos.buscar_atleta(id_atleta)
            
            if atleta and atleta.disciplina == evento.disciplina:
                #Solicita el puntaje del atleta y registra la participación si todo es válido
                puntaje = Herramientas.solicitar_flotante("Ingrese el puntaje del atleta (0-10): ")
                juegos_olimpicos.registrar_participacion(id_evento, id_atleta, puntaje)
                print("Participación registrada correctamente.")
            else:
                print("La disciplina del atleta no coincide con la del evento o el atleta no existe.")
        else:
            print("No se encontró el evento seleccionado.")
    else:
        print("No hay eventos en la fecha ingresada.")


def cantidad_atletas_por_disciplina():
    disciplinas = juegos_olimpicos.lista_disciplinas
    cantidades = []
    
    for disciplina in disciplinas:
        contador = 0
        #Recorre la lista Atletas y con un contador se suma las Disciplinas de los mismos
        for atleta in juegos_olimpicos.lista_atletas:
            if atleta.disciplina == disciplina:
                contador += 1
        cantidades.append(contador)
    
    nombres_disciplinas = [disciplina.nombre for disciplina in disciplinas]
    
    plt.bar(nombres_disciplinas, cantidades) #Gráfico de barras con librería matplotlib
    plt.xlabel("Disciplinas")
    plt.ylabel("Cantidad de Atletas")
    plt.title("Cantidad de Atletas por Disciplina")
    plt.show()

#Inicio del Programa
cargar_datos()
mostrar_menu()
opcion = solicitar_opcion()

while opcion != 0:
    if opcion == 1:
        ingresar_atleta()
    elif opcion == 2:
        ingreso_disciplinas_desde_excel()
    elif opcion == 3:
        asignar_disciplina_atleta()
    elif opcion == 4:
        ingresar_evento()
    elif opcion == 5:
        registrar_participacion()
    elif opcion == 6:
        cantidad_atletas_por_disciplina()
    
    guardar_datos()
    input("Presione Enter para continuar...")
    system("cls")
    mostrar_menu()
    opcion = solicitar_opcion()